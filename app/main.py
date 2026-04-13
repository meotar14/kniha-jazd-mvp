import csv
from copy import copy
import json
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from typing import Literal

import httpx
from fastapi import Depends, FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, Response
from openpyxl import Workbook, load_workbook
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload

from . import generator, models, schemas
from .db import Base, engine, get_db
from .version import APP_CONTACTS, APP_VERSION

app = FastAPI(title="Kniha jazd API", version=APP_VERSION)
STATIC_INDEX = Path(__file__).parent / "static" / "index.html"
TRIPS_TEMPLATE_XLSX = Path(__file__).parent / "templates" / "kniha_jazd_template.xlsx"
MONTH_NAMES_SK = {
    1: "januar",
    2: "februar",
    3: "marec",
    4: "april",
    5: "maj",
    6: "jun",
    7: "jul",
    8: "august",
    9: "september",
    10: "oktober",
    11: "november",
    12: "december",
}
BACKUP_SECTIONS = ("settings", "drivers", "vehicles", "customers", "month_plans", "trips", "refuels", "holidays")


def run_lightweight_migrations() -> None:
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE vehicles ADD COLUMN IF NOT EXISTS tank_capacity_l FLOAT DEFAULT 50"))
        conn.execute(text("UPDATE vehicles SET tank_capacity_l = 50 WHERE tank_capacity_l IS NULL"))
        conn.execute(text("ALTER TABLE vehicles ALTER COLUMN tank_capacity_l SET NOT NULL"))
        conn.execute(text("ALTER TABLE refuels ADD COLUMN IF NOT EXISTS total_price_eur FLOAT"))
        conn.execute(text("ALTER TABLE refuels ADD COLUMN IF NOT EXISTS location_city VARCHAR(128)"))
        conn.execute(text("ALTER TABLE refuels ADD COLUMN IF NOT EXISTS is_foreign BOOLEAN DEFAULT FALSE"))
        conn.execute(text("UPDATE refuels SET is_foreign = FALSE WHERE is_foreign IS NULL"))
        conn.execute(text("ALTER TABLE refuels ALTER COLUMN is_foreign SET NOT NULL"))
        conn.execute(text("ALTER TABLE trips ADD COLUMN IF NOT EXISTS trip_end_date DATE"))
        conn.execute(text("ALTER TABLE customers ADD COLUMN IF NOT EXISTS active_for_generation BOOLEAN DEFAULT TRUE"))
        conn.execute(text("UPDATE customers SET active_for_generation = TRUE WHERE active_for_generation IS NULL"))
        conn.execute(text("ALTER TABLE customers ALTER COLUMN active_for_generation SET NOT NULL"))
        conn.execute(text("ALTER TABLE month_plans ADD COLUMN IF NOT EXISTS private_km_enabled BOOLEAN DEFAULT FALSE"))
        conn.execute(text("UPDATE month_plans SET private_km_enabled = FALSE WHERE private_km_enabled IS NULL"))
        conn.execute(text("ALTER TABLE month_plans ALTER COLUMN private_km_enabled SET NOT NULL"))
        conn.execute(text("ALTER TABLE month_plans ADD COLUMN IF NOT EXISTS private_km_ratio_percent FLOAT DEFAULT 10"))
        conn.execute(text("UPDATE month_plans SET private_km_ratio_percent = 10 WHERE private_km_ratio_percent IS NULL"))
        conn.execute(text("ALTER TABLE month_plans ALTER COLUMN private_km_ratio_percent SET NOT NULL"))
        conn.execute(text("ALTER TABLE vehicles ADD COLUMN IF NOT EXISTS default_driver_id INTEGER"))
        conn.execute(text("ALTER TABLE vehicles ADD COLUMN IF NOT EXISTS use_custom_customer_catalog BOOLEAN DEFAULT FALSE"))
        conn.execute(text("UPDATE vehicles SET use_custom_customer_catalog = FALSE WHERE use_custom_customer_catalog IS NULL"))
        conn.execute(text("ALTER TABLE vehicles ALTER COLUMN use_custom_customer_catalog SET NOT NULL"))
        conn.execute(text("ALTER TABLE trips ADD COLUMN IF NOT EXISTS is_private BOOLEAN DEFAULT FALSE"))
        conn.execute(text("UPDATE trips SET is_private = FALSE WHERE is_private IS NULL"))
        conn.execute(text("ALTER TABLE trips ALTER COLUMN is_private SET NOT NULL"))
        conn.execute(text("ALTER TABLE customers ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"))
        conn.execute(text("UPDATE customers SET created_at = CURRENT_TIMESTAMP WHERE created_at IS NULL"))
        conn.execute(text("ALTER TABLE customers ALTER COLUMN created_at SET NOT NULL"))
        conn.execute(text("ALTER TABLE customers ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"))
        conn.execute(text("UPDATE customers SET updated_at = CURRENT_TIMESTAMP WHERE updated_at IS NULL"))
        conn.execute(text("ALTER TABLE customers ALTER COLUMN updated_at SET NOT NULL"))
        conn.execute(text("ALTER TABLE customers ADD COLUMN IF NOT EXISTS vehicle_id INTEGER"))
        conn.execute(text("ALTER TABLE customers ADD COLUMN IF NOT EXISTS source_customer_id INTEGER"))


def ensure_settings_row(db: Session) -> models.AppSettings:
    row = db.query(models.AppSettings).filter(models.AppSettings.id == 1).first()
    if row:
        return row
    row = models.AppSettings(id=1, company_name="", company_ico="", company_logo_url=None, company_base_address=None)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@app.on_event("startup")
def on_startup() -> None:
    Base.metadata.create_all(bind=engine)
    run_lightweight_migrations()


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/meta")
def meta() -> dict:
    return {"version": APP_VERSION, "contacts": APP_CONTACTS}


@app.get("/ui", include_in_schema=False)
def ui() -> FileResponse:
    return FileResponse(STATIC_INDEX)


@app.get("/settings")
def get_settings(db: Session = Depends(get_db)) -> dict:
    return serialize_settings(ensure_settings_row(db))


@app.put("/settings")
def update_settings(payload: schemas.AppSettingsUpdate, db: Session = Depends(get_db)) -> dict:
    row = ensure_settings_row(db)
    row.company_name = payload.company_name.strip()
    row.company_ico = payload.company_ico.strip()
    row.company_logo_url = payload.company_logo_url.strip() if payload.company_logo_url else None
    row.company_base_address = payload.company_base_address.strip() if payload.company_base_address else None
    db.commit()
    db.refresh(row)
    return serialize_settings(row)


@app.get("/holidays")
def list_holidays(year: int | None = None, db: Session = Depends(get_db)) -> list[dict]:
    target_year = year or datetime.utcnow().year
    ensure_holidays_for_year(db, target_year)
    query = db.query(models.Holiday)
    if year is not None:
        query = query.filter(
            models.Holiday.holiday_date >= date(year, 1, 1),
            models.Holiday.holiday_date <= date(year, 12, 31),
        )
    rows = query.order_by(models.Holiday.holiday_date.asc(), models.Holiday.id.asc()).all()
    return [serialize_holiday(row) for row in rows]


@app.post("/holidays")
def create_holiday(payload: schemas.HolidayCreate, db: Session = Depends(get_db)) -> dict:
    existing = db.query(models.Holiday).filter(models.Holiday.holiday_date == payload.holiday_date).first()
    if existing:
        existing.name = payload.name
        db.commit()
        db.refresh(existing)
        return serialize_holiday(existing)
    row = models.Holiday(holiday_date=payload.holiday_date, name=payload.name)
    db.add(row)
    db.commit()
    db.refresh(row)
    return serialize_holiday(row)


@app.delete("/holidays/{holiday_id}")
def delete_holiday(holiday_id: int, db: Session = Depends(get_db)) -> dict:
    row = db.get(models.Holiday, holiday_id)
    if not row:
        raise HTTPException(status_code=404, detail="holiday not found")
    db.delete(row)
    db.commit()
    return {"deleted": True, "id": holiday_id}


@app.post("/settings/logo")
async def upload_logo(file: UploadFile = File(...), db: Session = Depends(get_db)) -> dict:
    allowed = {".png", ".jpg", ".jpeg", ".webp", ".svg"}
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in allowed:
        raise HTTPException(status_code=400, detail="unsupported logo format")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="empty logo file")

    uploads_dir = Path(__file__).parent / "static" / "uploads"
    uploads_dir.mkdir(parents=True, exist_ok=True)
    logo_path = uploads_dir / f"company_logo{suffix}"
    logo_path.write_bytes(content)

    row = ensure_settings_row(db)
    row.company_logo_url = f"/assets/company_logo{suffix}"
    db.commit()
    db.refresh(row)
    return serialize_settings(row)


@app.get("/assets/{filename}", include_in_schema=False)
def assets(filename: str) -> FileResponse:
    safe_name = Path(filename).name
    static_root = Path(__file__).parent / "static"
    candidates = [
        static_root / "uploads" / safe_name,
        static_root / "branding" / safe_name,
    ]
    for asset_path in candidates:
        if asset_path.exists():
            return FileResponse(asset_path)
    raise HTTPException(status_code=404, detail="asset not found")


def _parse_backup_date(value: str | None) -> date | None:
    if not value:
        return None
    text = value.strip()
    if not text:
        return None
    if "." in text:
        day, month, year = text.split(".")
        return date(int(year), int(month), int(day))
    return date.fromisoformat(text)


def _easter_sunday(year: int) -> date:
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def get_slovak_public_holidays(year: int) -> list[tuple[date, str]]:
    easter = _easter_sunday(year)
    return [
        (date(year, 1, 1), "Deň vzniku Slovenskej republiky"),
        (date(year, 1, 6), "Zjavenie Pána"),
        (easter - timedelta(days=2), "Veľký piatok"),
        (easter + timedelta(days=1), "Veľkonočný pondelok"),
        (date(year, 5, 1), "Sviatok práce"),
        (date(year, 5, 8), "Deň víťazstva nad fašizmom"),
        (date(year, 7, 5), "Sviatok svätého Cyrila a Metoda"),
        (date(year, 8, 29), "Výročie SNP"),
        (date(year, 9, 1), "Deň Ústavy Slovenskej republiky"),
        (date(year, 9, 15), "Sedembolestná Panna Mária"),
        (date(year, 11, 1), "Sviatok všetkých svätých"),
        (date(year, 11, 17), "Deň boja za slobodu a demokraciu"),
        (date(year, 12, 24), "Štedrý deň"),
        (date(year, 12, 25), "Prvý sviatok vianočný"),
        (date(year, 12, 26), "Druhý sviatok vianočný"),
    ]


def ensure_holidays_for_year(db: Session, year: int) -> None:
    existing_dates = {
        row.holiday_date
        for row in db.query(models.Holiday).filter(
            models.Holiday.holiday_date >= date(year, 1, 1),
            models.Holiday.holiday_date <= date(year, 12, 31),
        )
    }
    inserted = False
    for holiday_date, name in get_slovak_public_holidays(year):
        if holiday_date in existing_dates:
            continue
        db.add(models.Holiday(holiday_date=holiday_date, name=name))
        inserted = True
    if inserted:
        db.commit()


def _reset_postgres_sequences(db: Session) -> None:
    bind = db.get_bind()
    if not bind or bind.dialect.name != "postgresql":
        return
    table_names = ["vehicles", "drivers", "customers", "month_plans", "trips", "refuels", "holidays"]
    for table in table_names:
        db.execute(
            text(
                f"""
                SELECT setval(
                    pg_get_serial_sequence('{table}', 'id'),
                    COALESCE((SELECT MAX(id) FROM {table}), 1),
                    COALESCE((SELECT MAX(id) IS NOT NULL FROM {table}), false)
                )
                """
            )
        )


def _normalize_backup_sections(values: list[str] | None) -> list[str]:
    if not values:
        return list(BACKUP_SECTIONS)
    sections: list[str] = []
    for raw in values:
        for part in str(raw).split(","):
            section = part.strip()
            if section and section in BACKUP_SECTIONS and section not in sections:
                sections.append(section)
    return sections or list(BACKUP_SECTIONS)


def _normalize_customer_key(name: str, address: str) -> tuple[str, str]:
    return (name.strip().casefold(), address.strip().casefold())


def _build_backup_payload(db: Session, sections: list[str]) -> dict:
    payload: dict = {
        "app": "kniha-jazd-mvp",
        "exported_at": datetime.utcnow().isoformat() + "Z",
        "meta": {"sections": sections, "app_version": APP_VERSION},
    }
    if "settings" in sections:
        payload["settings"] = serialize_settings(ensure_settings_row(db))
    if "vehicles" in sections:
        payload["vehicles"] = [
            {
                "id": v.id,
                "plate_number": v.plate_number,
                "model": v.model,
                "expected_consumption_l_per_100km": v.expected_consumption_l_per_100km,
                "tank_capacity_l": v.tank_capacity_l,
                "default_driver_id": v.default_driver_id,
                "use_custom_customer_catalog": v.use_custom_customer_catalog,
            }
            for v in db.query(models.Vehicle).order_by(models.Vehicle.id.asc()).all()
        ]
    if "drivers" in sections:
        payload["drivers"] = [
            {"id": d.id, "full_name": d.full_name, "license_number": d.license_number}
            for d in db.query(models.Driver).order_by(models.Driver.id.asc()).all()
        ]
    if "customers" in sections:
        payload["customers"] = [
            {
                "id": c.id,
                "name": c.name,
                "address": c.address,
                "distance_from_base_km": c.distance_from_base_km,
                "active_for_generation": c.active_for_generation,
                "vehicle_id": c.vehicle_id,
                "source_customer_id": c.source_customer_id,
                "created_at": c.created_at.isoformat() if c.created_at else None,
                "updated_at": c.updated_at.isoformat() if c.updated_at else None,
            }
            for c in db.query(models.Customer).order_by(models.Customer.id.asc()).all()
        ]
    if "month_plans" in sections:
        payload["month_plans"] = [
            {
                "id": p.id,
                "vehicle_id": p.vehicle_id,
                "driver_id": p.driver_id,
                "year": p.year,
                "month": p.month,
                "base_address": p.base_address,
                "start_odometer_km": p.start_odometer_km,
                "end_odometer_km": p.end_odometer_km,
                "private_km_enabled": p.private_km_enabled,
                "private_km_ratio_percent": p.private_km_ratio_percent,
            }
            for p in db.query(models.MonthPlan).order_by(models.MonthPlan.id.asc()).all()
        ]
    if "trips" in sections:
        payload["trips"] = [
            {
                "id": t.id,
                "month_plan_id": t.month_plan_id,
                "trip_date": t.trip_date.isoformat(),
                "trip_end_date": t.trip_end_date.isoformat() if t.trip_end_date else None,
                "customer_id": t.customer_id,
                "start_address": t.start_address,
                "end_address": t.end_address,
                "distance_km": t.distance_km,
                "generated": t.generated,
                "is_private": t.is_private,
                "note": t.note,
            }
            for t in db.query(models.Trip).order_by(models.Trip.id.asc()).all()
        ]
    if "refuels" in sections:
        payload["refuels"] = [
            {
                "id": r.id,
                "month_plan_id": r.month_plan_id,
                "refuel_date": r.refuel_date.isoformat(),
                "liters": r.liters,
                "odometer_km": r.odometer_km,
                "total_price_eur": r.total_price_eur,
                "location_city": r.location_city,
                "is_foreign": r.is_foreign,
            }
            for r in db.query(models.Refuel).order_by(models.Refuel.id.asc()).all()
        ]
    if "holidays" in sections:
        payload["holidays"] = [
            {"id": h.id, "holiday_date": h.holiday_date.isoformat(), "name": h.name}
            for h in db.query(models.Holiday).order_by(models.Holiday.holiday_date.asc(), models.Holiday.id.asc()).all()
        ]
    return payload


def _full_restore_backup(db: Session, payload: dict) -> dict:
    db.query(models.Holiday).delete()
    db.query(models.Refuel).delete()
    db.query(models.Trip).delete()
    db.query(models.MonthPlan).delete()
    db.query(models.Customer).delete()
    db.query(models.Driver).delete()
    db.query(models.Vehicle).delete()
    db.query(models.AppSettings).delete()

    s = payload.get("settings") or {}
    db.add(
        models.AppSettings(
            id=1,
            company_name=(s.get("company_name") or "").strip(),
            company_ico=(s.get("company_ico") or "").strip(),
            company_logo_url=s.get("company_logo_url"),
            company_base_address=s.get("company_base_address"),
        )
    )

    for d in payload.get("drivers", []):
        db.add(models.Driver(id=int(d["id"]), full_name=d["full_name"], license_number=d["license_number"]))
    for v in payload.get("vehicles", []):
        db.add(
            models.Vehicle(
                id=int(v["id"]),
                plate_number=v["plate_number"],
                model=v["model"],
                expected_consumption_l_per_100km=float(v["expected_consumption_l_per_100km"]),
                tank_capacity_l=float(v.get("tank_capacity_l", 50)),
                default_driver_id=int(v["default_driver_id"]) if v.get("default_driver_id") else None,
                use_custom_customer_catalog=bool(v.get("use_custom_customer_catalog", False)),
            )
        )
    for c in payload.get("customers", []):
        db.add(
            models.Customer(
                id=int(c["id"]),
                name=c["name"],
                address=c["address"],
                distance_from_base_km=float(c["distance_from_base_km"]),
                active_for_generation=bool(c.get("active_for_generation", True)),
                vehicle_id=int(c["vehicle_id"]) if c.get("vehicle_id") is not None else None,
                source_customer_id=int(c["source_customer_id"]) if c.get("source_customer_id") is not None else None,
                created_at=datetime.fromisoformat(c["created_at"]) if c.get("created_at") else datetime.utcnow(),
                updated_at=datetime.fromisoformat(c["updated_at"]) if c.get("updated_at") else datetime.utcnow(),
            )
        )
    for p in payload.get("month_plans", []):
        db.add(
            models.MonthPlan(
                id=int(p["id"]),
                vehicle_id=int(p["vehicle_id"]),
                driver_id=int(p["driver_id"]),
                year=int(p["year"]),
                month=int(p["month"]),
                base_address=p["base_address"],
                start_odometer_km=int(p["start_odometer_km"]),
                end_odometer_km=int(p["end_odometer_km"]),
                private_km_enabled=bool(p.get("private_km_enabled", False)),
                private_km_ratio_percent=float(p.get("private_km_ratio_percent", 10.0)),
            )
        )
    for t in payload.get("trips", []):
        db.add(
            models.Trip(
                id=int(t["id"]),
                month_plan_id=int(t["month_plan_id"]),
                trip_date=_parse_backup_date(t.get("trip_date")),
                trip_end_date=_parse_backup_date(t.get("trip_end_date")),
                customer_id=int(t["customer_id"]) if t.get("customer_id") is not None else None,
                start_address=t["start_address"],
                end_address=t["end_address"],
                distance_km=float(t["distance_km"]),
                generated=bool(t.get("generated", False)),
                is_private=bool(t.get("is_private", False)),
                note=t.get("note"),
            )
        )
    for r in payload.get("refuels", []):
        db.add(
            models.Refuel(
                id=int(r["id"]),
                month_plan_id=int(r["month_plan_id"]),
                refuel_date=_parse_backup_date(r.get("refuel_date")),
                liters=float(r["liters"]),
                odometer_km=int(r["odometer_km"]) if r.get("odometer_km") is not None else None,
                total_price_eur=float(r["total_price_eur"]) if r.get("total_price_eur") is not None else None,
                location_city=r.get("location_city"),
                is_foreign=bool(r.get("is_foreign", False)),
            )
        )
    for h in payload.get("holidays", []):
        db.add(models.Holiday(id=int(h["id"]), holiday_date=_parse_backup_date(h.get("holiday_date")), name=h["name"]))
    db.flush()
    _reset_postgres_sequences(db)
    return {
        "mode": "replace_all",
        "vehicles": len(payload.get("vehicles", [])),
        "drivers": len(payload.get("drivers", [])),
        "customers": len(payload.get("customers", [])),
        "month_plans": len(payload.get("month_plans", [])),
        "trips": len(payload.get("trips", [])),
        "refuels": len(payload.get("refuels", [])),
        "holidays": len(payload.get("holidays", [])),
    }


def _merge_backup_sections(db: Session, payload: dict, sections: list[str], replace_existing: bool) -> dict:
    driver_id_map: dict[int, int] = {}
    vehicle_id_map: dict[int, int] = {}
    customer_id_map: dict[int, int] = {}
    month_plan_id_map: dict[int, int] = {}

    if "settings" in sections:
        s = payload.get("settings") or {}
        row = ensure_settings_row(db)
        row.company_name = (s.get("company_name") or "").strip()
        row.company_ico = (s.get("company_ico") or "").strip()
        row.company_logo_url = s.get("company_logo_url")
        row.company_base_address = s.get("company_base_address")

    if "drivers" in sections:
        existing_by_license = {d.license_number: d for d in db.query(models.Driver).all()}
        for d in payload.get("drivers", []):
            row = existing_by_license.get(d["license_number"])
            if not row:
                row = models.Driver(full_name=d["full_name"], license_number=d["license_number"])
                db.add(row)
                db.flush()
                existing_by_license[row.license_number] = row
            else:
                row.full_name = d["full_name"]
            driver_id_map[int(d["id"])] = row.id
    else:
        driver_id_map = {d.id: d.id for d in db.query(models.Driver).all()}

    if "vehicles" in sections:
        existing_by_plate = {v.plate_number: v for v in db.query(models.Vehicle).all()}
        for v in payload.get("vehicles", []):
            row = existing_by_plate.get(v["plate_number"])
            if not row:
                row = models.Vehicle(plate_number=v["plate_number"], model=v["model"], expected_consumption_l_per_100km=float(v["expected_consumption_l_per_100km"]), tank_capacity_l=float(v.get("tank_capacity_l", 50)))
                db.add(row)
                db.flush()
                existing_by_plate[row.plate_number] = row
            row.model = v["model"]
            row.expected_consumption_l_per_100km = float(v["expected_consumption_l_per_100km"])
            row.tank_capacity_l = float(v.get("tank_capacity_l", 50))
            default_driver_id = int(v["default_driver_id"]) if v.get("default_driver_id") else None
            row.default_driver_id = driver_id_map.get(default_driver_id) if default_driver_id else None
            row.use_custom_customer_catalog = bool(v.get("use_custom_customer_catalog", False))
            vehicle_id_map[int(v["id"])] = row.id
    else:
        vehicle_id_map = {v.id: v.id for v in db.query(models.Vehicle).all()}

    if "customers" in sections:
        existing_customers = db.query(models.Customer).all()
        existing_by_key = {(_normalize_customer_key(c.name, c.address), c.vehicle_id): c for c in existing_customers}
        for c in payload.get("customers", []):
            mapped_vehicle_id = vehicle_id_map.get(int(c["vehicle_id"])) if c.get("vehicle_id") is not None else None
            key = (_normalize_customer_key(c["name"], c["address"]), mapped_vehicle_id)
            row = existing_by_key.get(key)
            if not row:
                row = models.Customer(
                    name=c["name"],
                    address=c["address"],
                    distance_from_base_km=float(c["distance_from_base_km"]),
                    active_for_generation=bool(c.get("active_for_generation", True)),
                    vehicle_id=mapped_vehicle_id,
                )
                db.add(row)
                db.flush()
                existing_by_key[key] = row
            row.distance_from_base_km = float(c["distance_from_base_km"])
            row.active_for_generation = bool(c.get("active_for_generation", True))
            row.vehicle_id = mapped_vehicle_id
            row.created_at = datetime.fromisoformat(c["created_at"]) if c.get("created_at") else row.created_at
            row.updated_at = datetime.fromisoformat(c["updated_at"]) if c.get("updated_at") else datetime.utcnow()
            customer_id_map[int(c["id"])] = row.id
    else:
        customer_id_map = {c.id: c.id for c in db.query(models.Customer).all()}

    if "month_plans" in sections:
        existing_plans = db.query(models.MonthPlan).all()
        existing_by_key = {(p.vehicle_id, p.year, p.month): p for p in existing_plans}
        for p in payload.get("month_plans", []):
            vehicle_id = vehicle_id_map.get(int(p["vehicle_id"]))
            driver_id = driver_id_map.get(int(p["driver_id"]))
            if not vehicle_id or not driver_id:
                continue
            key = (vehicle_id, int(p["year"]), int(p["month"]))
            row = existing_by_key.get(key)
            if not row:
                row = models.MonthPlan(
                    vehicle_id=vehicle_id,
                    driver_id=driver_id,
                    year=int(p["year"]),
                    month=int(p["month"]),
                    base_address=p["base_address"],
                    start_odometer_km=int(p["start_odometer_km"]),
                    end_odometer_km=int(p["end_odometer_km"]),
                    private_km_enabled=bool(p.get("private_km_enabled", False)),
                    private_km_ratio_percent=float(p.get("private_km_ratio_percent", 10.0)),
                )
                db.add(row)
                db.flush()
                existing_by_key[key] = row
            row.driver_id = driver_id
            row.base_address = p["base_address"]
            row.start_odometer_km = int(p["start_odometer_km"])
            row.end_odometer_km = int(p["end_odometer_km"])
            row.private_km_enabled = bool(p.get("private_km_enabled", False))
            row.private_km_ratio_percent = float(p.get("private_km_ratio_percent", 10.0))
            month_plan_id_map[int(p["id"])] = row.id
    else:
        month_plan_id_map = {p.id: p.id for p in db.query(models.MonthPlan).all()}

    if "holidays" in sections:
        existing_by_date = {h.holiday_date: h for h in db.query(models.Holiday).all()}
        for h in payload.get("holidays", []):
            holiday_date = _parse_backup_date(h.get("holiday_date"))
            row = existing_by_date.get(holiday_date)
            if not row:
                row = models.Holiday(holiday_date=holiday_date, name=h["name"])
                db.add(row)
                existing_by_date[holiday_date] = row
            else:
                row.name = h["name"]

    if "refuels" in sections and replace_existing:
        imported_plan_ids = [month_plan_id_map.get(int(r["month_plan_id"])) for r in payload.get("refuels", [])]
        imported_plan_ids = [plan_id for plan_id in imported_plan_ids if plan_id]
        if imported_plan_ids:
            db.query(models.Refuel).filter(models.Refuel.month_plan_id.in_(imported_plan_ids)).delete(synchronize_session=False)

    if "trips" in sections and replace_existing:
        imported_plan_ids = [month_plan_id_map.get(int(t["month_plan_id"])) for t in payload.get("trips", [])]
        imported_plan_ids = [plan_id for plan_id in imported_plan_ids if plan_id]
        if imported_plan_ids:
            db.query(models.Trip).filter(models.Trip.month_plan_id.in_(imported_plan_ids)).delete(synchronize_session=False)

    imported_trips = 0
    if "trips" in sections:
        for t in payload.get("trips", []):
            plan_id = month_plan_id_map.get(int(t["month_plan_id"]))
            if not plan_id:
                continue
            customer_id = customer_id_map.get(int(t["customer_id"])) if t.get("customer_id") is not None else None
            trip_date = _parse_backup_date(t.get("trip_date"))
            trip_end_date = _parse_backup_date(t.get("trip_end_date"))
            query = db.query(models.Trip).filter(
                models.Trip.month_plan_id == plan_id,
                models.Trip.trip_date == trip_date,
                models.Trip.trip_end_date == trip_end_date,
                models.Trip.distance_km == float(t["distance_km"]),
                models.Trip.generated == bool(t.get("generated", False)),
                models.Trip.is_private == bool(t.get("is_private", False)),
                models.Trip.note == t.get("note"),
            )
            row = None if replace_existing else query.first()
            if not row:
                row = models.Trip(
                    month_plan_id=plan_id,
                    trip_date=trip_date,
                    trip_end_date=trip_end_date,
                    customer_id=customer_id,
                    start_address=t["start_address"],
                    end_address=t["end_address"],
                    distance_km=float(t["distance_km"]),
                    generated=bool(t.get("generated", False)),
                    is_private=bool(t.get("is_private", False)),
                    note=t.get("note"),
                )
                db.add(row)
            else:
                row.customer_id = customer_id
                row.start_address = t["start_address"]
                row.end_address = t["end_address"]
            imported_trips += 1

    imported_refuels = 0
    if "refuels" in sections:
        for r in payload.get("refuels", []):
            plan_id = month_plan_id_map.get(int(r["month_plan_id"]))
            if not plan_id:
                continue
            refuel_date = _parse_backup_date(r.get("refuel_date"))
            query = db.query(models.Refuel).filter(
                models.Refuel.month_plan_id == plan_id,
                models.Refuel.refuel_date == refuel_date,
                models.Refuel.liters == float(r["liters"]),
                models.Refuel.odometer_km == (int(r["odometer_km"]) if r.get("odometer_km") is not None else None),
            )
            row = None if replace_existing else query.first()
            if not row:
                row = models.Refuel(
                    month_plan_id=plan_id,
                    refuel_date=refuel_date,
                    liters=float(r["liters"]),
                    odometer_km=int(r["odometer_km"]) if r.get("odometer_km") is not None else None,
                    total_price_eur=float(r["total_price_eur"]) if r.get("total_price_eur") is not None else None,
                    location_city=r.get("location_city"),
                    is_foreign=bool(r.get("is_foreign", False)),
                )
                db.add(row)
            else:
                row.total_price_eur = float(r["total_price_eur"]) if r.get("total_price_eur") is not None else None
                row.location_city = r.get("location_city")
                row.is_foreign = bool(r.get("is_foreign", False))
            imported_refuels += 1

    db.flush()
    _reset_postgres_sequences(db)
    return {
        "mode": "replace_selected" if replace_existing else "merge",
        "vehicles": len(payload.get("vehicles", [])) if "vehicles" in sections else 0,
        "drivers": len(payload.get("drivers", [])) if "drivers" in sections else 0,
        "customers": len(payload.get("customers", [])) if "customers" in sections else 0,
        "month_plans": len(payload.get("month_plans", [])) if "month_plans" in sections else 0,
        "trips": imported_trips,
        "refuels": imported_refuels,
        "holidays": len(payload.get("holidays", [])) if "holidays" in sections else 0,
        "sections": sections,
    }


@app.get("/backup/export")
def export_backup_json(sections: list[str] = Query(default=[]), db: Session = Depends(get_db)) -> Response:
    payload = _build_backup_payload(db, _normalize_backup_sections(sections))
    backup_name = f"kniha_jazd_backup_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
    return Response(
        content=json.dumps(payload, ensure_ascii=False, indent=2),
        media_type="application/json; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{backup_name}"'},
    )


@app.post("/backup/import")
async def import_backup_json(
    file: UploadFile = File(...),
    replace_existing: bool = Form(True),
    sections: list[str] = Form(default=[]),
    db: Session = Depends(get_db),
) -> dict:
    if not file.filename.lower().endswith(".json"):
        raise HTTPException(status_code=400, detail="only .json backup files are supported")
    try:
        payload = json.loads((await file.read()).decode("utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="invalid backup json") from exc

    payload_sections = payload.get("meta", {}).get("sections")
    selected_sections = _normalize_backup_sections(sections or payload_sections)
    if not any(section in payload for section in selected_sections):
        raise HTTPException(status_code=400, detail="backup json does not contain selected sections")

    try:
        is_full_restore = replace_existing and set(selected_sections) == set(BACKUP_SECTIONS)
        result = _full_restore_backup(db, payload) if is_full_restore else _merge_backup_sections(db, payload, selected_sections, replace_existing)
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"backup import failed: {exc}") from exc

    return {
        "imported": True,
        **result,
    }


def geocode_address(address: str) -> tuple[float, float]:
    with httpx.Client(timeout=15) as client:
        response = client.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": address, "format": "json", "limit": 1},
            headers={"User-Agent": "kniha-jazd-mvp/0.2.0"},
        )
        response.raise_for_status()
        data = response.json()
        if not data:
            raise HTTPException(status_code=404, detail=f"address not found: {address}")
        return float(data[0]["lat"]), float(data[0]["lon"])


@app.get("/distance-estimate")
def distance_estimate(base_address: str, destination_address: str) -> dict:
    start_lat, start_lon = geocode_address(base_address)
    end_lat, end_lon = geocode_address(destination_address)
    with httpx.Client(timeout=15) as client:
        route_response = client.get(
            f"https://router.project-osrm.org/route/v1/driving/{start_lon},{start_lat};{end_lon},{end_lat}",
            params={"overview": "false"},
            headers={"User-Agent": "kniha-jazd-mvp/0.2.0"},
        )
        route_response.raise_for_status()
        route_data = route_response.json()
        routes = route_data.get("routes") or []
        if not routes:
            raise HTTPException(status_code=404, detail="route not found")
        one_way_km = routes[0]["distance"] / 1000.0
        return {
            "one_way_km": round(one_way_km, 1),
            "roundtrip_km": round(one_way_km * 2, 1),
        }


def serialize_vehicle(vehicle: models.Vehicle) -> dict:
    return {
        "id": vehicle.id,
        "plate_number": vehicle.plate_number,
        "model": vehicle.model,
        "expected_consumption_l_per_100km": vehicle.expected_consumption_l_per_100km,
        "tank_capacity_l": vehicle.tank_capacity_l,
        "default_driver_id": vehicle.default_driver_id,
        "default_driver_name": vehicle.default_driver.full_name if vehicle.default_driver else None,
        "use_custom_customer_catalog": vehicle.use_custom_customer_catalog,
    }


def serialize_driver(driver: models.Driver) -> dict:
    return {
        "id": driver.id,
        "full_name": driver.full_name,
        "license_number": driver.license_number,
    }


def serialize_customer(customer: models.Customer) -> dict:
    catalog_name = "Globálny"
    if customer.vehicle_id and customer.vehicle:
        catalog_name = f"{customer.vehicle.plate_number} | vlastný"
    return {
        "id": customer.id,
        "name": customer.name,
        "address": customer.address,
        "distance_from_base_km": customer.distance_from_base_km,
        "active_for_generation": customer.active_for_generation,
        "vehicle_id": customer.vehicle_id,
        "source_customer_id": customer.source_customer_id,
        "catalog_name": catalog_name,
        "created_at": customer.created_at.isoformat() if customer.created_at else None,
        "updated_at": customer.updated_at.isoformat() if customer.updated_at else None,
    }


def serialize_holiday(row: models.Holiday) -> dict:
    return {
        "id": row.id,
        "holiday_date": row.holiday_date.isoformat(),
        "name": row.name,
    }


def ensure_vehicle_customer_catalog(db: Session, vehicle: models.Vehicle) -> None:
    if not vehicle.use_custom_customer_catalog:
        return
    existing = db.query(models.Customer).filter(models.Customer.vehicle_id == vehicle.id).first()
    if existing:
        return
    base_customers = db.query(models.Customer).filter(models.Customer.vehicle_id.is_(None)).order_by(models.Customer.id.asc()).all()
    for customer in base_customers:
        db.add(
            models.Customer(
                name=customer.name,
                address=customer.address,
                distance_from_base_km=customer.distance_from_base_km,
                active_for_generation=customer.active_for_generation,
                vehicle_id=vehicle.id,
                source_customer_id=customer.id,
                created_at=customer.created_at,
                updated_at=customer.updated_at,
            )
        )
    db.flush()


def serialize_month_plan(month_plan: models.MonthPlan) -> dict:
    month_km = round(sum(t.distance_km for t in month_plan.trips), 1)
    private_km = round(sum(t.distance_km for t in month_plan.trips if t.is_private), 1)
    service_target_km = generator.plan_service_target_km(month_plan)
    return {
        "id": month_plan.id,
        "vehicle_id": month_plan.vehicle_id,
        "vehicle_plate_number": month_plan.vehicle.plate_number,
        "vehicle_model": month_plan.vehicle.model,
        "driver_id": month_plan.driver_id,
        "driver_name": month_plan.driver.full_name,
        "year": month_plan.year,
        "month": month_plan.month,
        "month_name": MONTH_NAMES_SK.get(month_plan.month, str(month_plan.month)),
        "month_km": month_km,
        "base_address": month_plan.base_address,
        "start_odometer_km": month_plan.start_odometer_km,
        "end_odometer_km": month_plan.end_odometer_km,
        "private_km_enabled": month_plan.private_km_enabled,
        "private_km_ratio_percent": month_plan.private_km_ratio_percent,
        "hidden_private_km": private_km,
        "service_target_km": service_target_km,
    }


def serialize_refuel(refuel: models.Refuel) -> dict:
    return {
        "id": refuel.id,
        "month_plan_id": refuel.month_plan_id,
        "refuel_date": refuel.refuel_date.isoformat(),
        "liters": refuel.liters,
        "odometer_km": refuel.odometer_km,
        "total_price_eur": refuel.total_price_eur,
        "location_city": refuel.location_city,
        "is_foreign": refuel.is_foreign,
    }


def serialize_trip(trip: models.Trip) -> dict:
    month_name = None
    if trip.month_plan:
        month_name = MONTH_NAMES_SK.get(trip.month_plan.month, str(trip.month_plan.month))
    start_address = "" if trip.is_private else trip.start_address
    end_address = "" if trip.is_private else trip.end_address
    return {
        "id": trip.id,
        "month_plan_id": trip.month_plan_id,
        "month_name": month_name,
        "trip_date": trip.trip_date.isoformat(),
        "trip_end_date": trip.trip_end_date.isoformat() if trip.trip_end_date else None,
        "customer_id": trip.customer_id,
        "customer_name": trip.customer.name if trip.customer else None,
        "start_address": start_address,
        "end_address": end_address,
        "distance_km": trip.distance_km,
        "generated": trip.generated,
        "is_private": trip.is_private,
        "note": trip.note,
    }


def validate_trip_range(month_plan: models.MonthPlan, trip_date, trip_end_date) -> None:
    end_date = trip_end_date or trip_date
    if end_date < trip_date:
        raise HTTPException(status_code=400, detail="trip_end_date must be >= trip_date")
    if trip_date.year != month_plan.year or trip_date.month != month_plan.month:
        raise HTTPException(status_code=400, detail="trip_date must be inside selected month plan")
    if end_date.year != month_plan.year or end_date.month != month_plan.month:
        raise HTTPException(status_code=400, detail="trip_end_date must be inside selected month plan")


def serialize_settings(row: models.AppSettings) -> dict:
    logo_url = row.company_logo_url
    if logo_url and "/assets/company-logo" in logo_url:
        logo_url = logo_url.replace("/assets/company-logo", "/assets/company_logo")
    if not logo_url:
        logo_url = "/assets/airo_default_logo.png"
    return {
        "company_name": row.company_name,
        "company_ico": row.company_ico,
        "company_logo_url": logo_url,
        "company_base_address": row.company_base_address,
        "app_version": APP_VERSION,
        "support_contacts": APP_CONTACTS,
    }


def trip_purpose_label(trip: models.Trip) -> str:
    if trip.is_private:
        return "Sukromna jazda"
    if trip.note:
        note = trip.note.strip()
        normalized_note = note.lower()
        if not trip.generated and note:
            return note
        if "auto-generated" not in normalized_note and "automaticky generovana jazda" not in normalized_note:
            return note
    if trip.customer:
        return f"Cesta k zakaznikovi: {trip.customer.name}"
    return "Sluzobna cesta"


def format_date_sk(date_iso: str) -> str:
    parts = date_iso.split("-")
    if len(parts) != 3:
        return date_iso
    return f"{parts[2]}.{parts[1]}.{parts[0]}"


def build_trip_odometer_rows(month_plan: models.MonthPlan, trips: list[models.Trip]) -> list[dict]:
    rows: list[dict] = []
    odometer = float(month_plan.start_odometer_km)
    for trip in sorted(trips, key=lambda t: (t.trip_date, t.id)):
        start_km = round(odometer, 1)
        end_km = round(start_km + trip.distance_km, 1)
        odometer = end_km
        rows.append(
            {
                "trip": trip,
                "odometer_start_km": start_km,
                "odometer_end_km": end_km,
            }
        )
    return rows


def render_trip_export_csv(rows: list[dict]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "Meno vodica",
            "SPZ",
            "Datum cesty",
            "Km na zaciatku",
            "Km na konci",
            "Ucel cesty",
            "Start adresa",
            "Ciel adresa",
            "Vzdialenost km",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                row["driver_name"],
                row["plate_number"],
                row["trip_date"],
                row["odometer_start_km"],
                row["odometer_end_km"],
                row["purpose"],
                row["start_address"],
                row["end_address"],
                row["distance_km"],
            ]
        )
    return buffer.getvalue()


def render_trip_export_xlsx(rows: list[dict]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Kniha jazd"
    headers = [
        "Meno vodica",
        "SPZ",
        "Datum cesty",
        "Km na zaciatku",
        "Km na konci",
        "Ucel cesty",
        "Start adresa",
        "Ciel adresa",
        "Vzdialenost km",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row["driver_name"],
                row["plate_number"],
                row["trip_date"],
                row["odometer_start_km"],
                row["odometer_end_km"],
                row["purpose"],
                row["start_address"],
                row["end_address"],
                row["distance_km"],
            ]
        )
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _resolve_template_sheet(workbook, month: int, year: int):
    preferred_sheet = f"{month:02d}_{year}"
    fallback_sheet = f"{month:02d}_2026"
    if preferred_sheet in workbook.sheetnames:
        return workbook[preferred_sheet]
    elif fallback_sheet in workbook.sheetnames:
        return workbook[fallback_sheet]
    raise HTTPException(status_code=400, detail=f"template sheet for month {month} not found")


def _fill_template_month_sheet(
    sheet,
    month_plan: models.MonthPlan,
    trips: list[models.Trip],
    company_name: str | None,
) -> None:

    data_start_row = 7
    base_capacity = 34
    summary_base_row = 41
    used_rows = len(trips)
    extra_rows = max(0, used_rows - base_capacity)

    month_name_display = MONTH_NAMES_SK.get(month_plan.month, str(month_plan.month)).capitalize()
    sheet["D2"] = month_plan.driver.full_name
    sheet["J2"] = company_name or "Spolocnost"
    sheet["A3"] = f"ŠPZ : {month_plan.vehicle.plate_number}"
    sheet["A4"] = f"Značka vozidla: {month_plan.vehicle.model}"
    sheet["E4"] = month_plan.vehicle.expected_consumption_l_per_100km
    sheet["J3"] = month_name_display
    sheet["J4"] = month_plan.year

    # For bigger exports, push summary rows down and clone row styles so layout stays usable.
    if extra_rows > 0:
        sheet.insert_rows(summary_base_row, amount=extra_rows)
        for i in range(extra_rows):
            src_row = data_start_row + base_capacity - 1
            dst_row = src_row + 1 + i
            for col in range(1, 12):  # A..K
                src_cell = sheet.cell(row=src_row, column=col)
                dst_cell = sheet.cell(row=dst_row, column=col)
                dst_cell._style = copy(src_cell._style)
                if src_cell.has_style and src_cell.number_format:
                    dst_cell.number_format = src_cell.number_format
            if src_row in sheet.row_dimensions:
                sheet.row_dimensions[dst_row].height = sheet.row_dimensions[src_row].height

    max_trip_rows = max(base_capacity, used_rows)
    # Clean trip table cells to avoid stale values from template.
    for row_no in range(data_start_row, data_start_row + max_trip_rows):
        for col in range(1, 12):  # A..K
            sheet.cell(row=row_no, column=col).value = None

    odometer_rows = build_trip_odometer_rows(month_plan, trips)
    for index, trip_row in enumerate(odometer_rows, start=1):
        trip = trip_row["trip"]
        row_no = data_start_row + index - 1
        start_km = trip_row["odometer_start_km"]
        end_km = trip_row["odometer_end_km"]

        sheet[f"A{row_no}"] = index
        sheet[f"B{row_no}"] = format_date_sk(trip.trip_date.isoformat())
        sheet[f"C{row_no}"] = "" if trip.is_private else trip.start_address
        sheet[f"D{row_no}"] = "" if trip.is_private else trip.end_address
        sheet[f"E{row_no}"] = ""
        sheet[f"F{row_no}"] = ""
        sheet[f"G{row_no}"] = start_km
        sheet[f"H{row_no}"] = end_km
        sheet[f"I{row_no}"] = round(trip.distance_km, 1)
        sheet[f"J{row_no}"] = month_plan.driver.full_name
        sheet[f"K{row_no}"] = trip_purpose_label(trip)

    summary_row = summary_base_row + extra_rows
    service_km = round(sum(t.distance_km for t in trips), 1)
    final_odometer = float(month_plan.end_odometer_km)
    sheet[f"I{summary_row}"] = service_km
    sheet[f"D{summary_row + 1}"] = month_plan.start_odometer_km
    sheet[f"D{summary_row + 2}"] = final_odometer
    sheet[f"D{summary_row + 3}"] = round(final_odometer - month_plan.start_odometer_km, 1)


def render_template_trip_export_xlsx(
    month_plan: models.MonthPlan,
    trips: list[models.Trip],
    company_name: str | None,
) -> bytes:
    if not TRIPS_TEMPLATE_XLSX.exists():
        raise HTTPException(status_code=500, detail="xlsx template not found on server")

    workbook = load_workbook(TRIPS_TEMPLATE_XLSX)
    sheet = _resolve_template_sheet(workbook, month_plan.month, month_plan.year)
    _fill_template_month_sheet(sheet, month_plan, trips, company_name)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_export_rows_for_month_plan(month_plan: models.MonthPlan, trips: list[models.Trip]) -> list[dict]:
    rows: list[dict] = []
    for trip_row in build_trip_odometer_rows(month_plan, trips):
        trip = trip_row["trip"]
        rows.append(
            {
                "driver_name": month_plan.driver.full_name,
                "plate_number": month_plan.vehicle.plate_number,
                "trip_date": format_date_sk(trip.trip_date.isoformat()),
                "odometer_start_km": trip_row["odometer_start_km"],
                "odometer_end_km": trip_row["odometer_end_km"],
                "purpose": trip_purpose_label(trip),
                "start_address": "" if trip.is_private else trip.start_address,
                "end_address": "" if trip.is_private else trip.end_address,
                "distance_km": trip.distance_km,
            }
        )
    return rows


def build_export_rows_for_mixed_trips(trips: list[models.Trip]) -> list[dict]:
    rows: list[dict] = []
    trips_by_plan: dict[int, list[models.Trip]] = {}
    for trip in trips:
        trips_by_plan.setdefault(trip.month_plan_id, []).append(trip)
    for plan_id, grouped_trips in trips_by_plan.items():
        month_plan = grouped_trips[0].month_plan
        for trip_row in build_trip_odometer_rows(month_plan, grouped_trips):
            trip = trip_row["trip"]
            rows.append(
                {
                    "driver_name": month_plan.driver.full_name,
                    "plate_number": month_plan.vehicle.plate_number,
                    "trip_date": format_date_sk(trip.trip_date.isoformat()),
                    "odometer_start_km": trip_row["odometer_start_km"],
                    "odometer_end_km": trip_row["odometer_end_km"],
                    "purpose": trip_purpose_label(trip),
                    "start_address": "" if trip.is_private else trip.start_address,
                    "end_address": "" if trip.is_private else trip.end_address,
                    "distance_km": trip.distance_km,
                }
            )
    return rows


@app.get("/vehicles")
def list_vehicles(db: Session = Depends(get_db)) -> list[dict]:
    rows = db.query(models.Vehicle).options(joinedload(models.Vehicle.default_driver)).order_by(models.Vehicle.id.asc()).all()
    return [serialize_vehicle(r) for r in rows]


@app.get("/vehicles/consumption-summary")
def vehicle_consumption_summary(db: Session = Depends(get_db)) -> list[dict]:
    vehicles = (
        db.query(models.Vehicle)
        .options(
            joinedload(models.Vehicle.month_plans).joinedload(models.MonthPlan.trips),
            joinedload(models.Vehicle.month_plans).joinedload(models.MonthPlan.refuels),
        )
        .order_by(models.Vehicle.id.asc())
        .all()
    )
    rows: list[dict] = []
    for vehicle in vehicles:
        total_km = round(sum(t.distance_km for p in vehicle.month_plans for t in p.trips), 1)
        total_refueled_l = round(sum(r.liters for p in vehicle.month_plans for r in p.refuels), 1)
        avg = round((total_refueled_l / total_km) * 100.0, 2) if total_km > 0 else None
        rows.append(
            {
                "vehicle_id": vehicle.id,
                "plate_number": vehicle.plate_number,
                "model": vehicle.model,
                "plans_count": len(vehicle.month_plans),
                "total_km": total_km,
                "total_refueled_l": total_refueled_l,
                "average_consumption_l_per_100km": avg,
            }
        )
    return rows


@app.post("/vehicles")
def create_vehicle(payload: schemas.VehicleCreate, db: Session = Depends(get_db)) -> dict:
    if payload.default_driver_id and not db.get(models.Driver, payload.default_driver_id):
        raise HTTPException(status_code=404, detail="default driver not found")
    vehicle = models.Vehicle(**payload.model_dump())
    db.add(vehicle)
    db.commit()
    db.refresh(vehicle)
    ensure_vehicle_customer_catalog(db, vehicle)
    db.commit()
    return {"id": vehicle.id, "plate_number": vehicle.plate_number}


@app.put("/vehicles/{vehicle_id}")
def update_vehicle(vehicle_id: int, payload: schemas.VehicleUpdate, db: Session = Depends(get_db)) -> dict:
    vehicle = db.get(models.Vehicle, vehicle_id)
    if not vehicle:
        raise HTTPException(status_code=404, detail="vehicle not found")
    if payload.default_driver_id and not db.get(models.Driver, payload.default_driver_id):
        raise HTTPException(status_code=404, detail="default driver not found")
    for key, value in payload.model_dump().items():
        setattr(vehicle, key, value)
    try:
        db.flush()
        ensure_vehicle_customer_catalog(db, vehicle)
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="vehicle update conflicts with existing data") from exc
    db.refresh(vehicle)
    return serialize_vehicle(vehicle)


@app.delete("/vehicles/{vehicle_id}")
def delete_vehicle(vehicle_id: int, db: Session = Depends(get_db)) -> dict:
    vehicle = db.get(models.Vehicle, vehicle_id)
    if not vehicle:
        raise HTTPException(status_code=404, detail="vehicle not found")
    if db.query(models.MonthPlan).filter(models.MonthPlan.vehicle_id == vehicle_id).first():
        raise HTTPException(status_code=409, detail="vehicle is used in month plans")
    db.query(models.Customer).filter(models.Customer.vehicle_id == vehicle_id).delete(synchronize_session=False)
    db.delete(vehicle)
    db.commit()
    return {"deleted": True, "id": vehicle_id}


@app.post("/drivers")
def create_driver(payload: schemas.DriverCreate, db: Session = Depends(get_db)) -> dict:
    driver = models.Driver(**payload.model_dump())
    db.add(driver)
    db.commit()
    db.refresh(driver)
    return {"id": driver.id, "full_name": driver.full_name}


@app.get("/drivers")
def list_drivers(db: Session = Depends(get_db)) -> list[dict]:
    rows = db.query(models.Driver).order_by(models.Driver.id.asc()).all()
    return [serialize_driver(r) for r in rows]


@app.put("/drivers/{driver_id}")
def update_driver(driver_id: int, payload: schemas.DriverUpdate, db: Session = Depends(get_db)) -> dict:
    driver = db.get(models.Driver, driver_id)
    if not driver:
        raise HTTPException(status_code=404, detail="driver not found")
    for key, value in payload.model_dump().items():
        setattr(driver, key, value)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="driver update conflicts with existing data") from exc
    db.refresh(driver)
    return serialize_driver(driver)


@app.delete("/drivers/{driver_id}")
def delete_driver(driver_id: int, db: Session = Depends(get_db)) -> dict:
    driver = db.get(models.Driver, driver_id)
    if not driver:
        raise HTTPException(status_code=404, detail="driver not found")
    if db.query(models.MonthPlan).filter(models.MonthPlan.driver_id == driver_id).first():
        raise HTTPException(status_code=409, detail="driver is used in month plans")
    if db.query(models.Vehicle).filter(models.Vehicle.default_driver_id == driver_id).first():
        raise HTTPException(status_code=409, detail="driver is used as default driver on a vehicle")
    db.delete(driver)
    db.commit()
    return {"deleted": True, "id": driver_id}


@app.post("/customers")
def create_customer(payload: schemas.CustomerCreate, db: Session = Depends(get_db)) -> dict:
    if payload.vehicle_id and not db.get(models.Vehicle, payload.vehicle_id):
        raise HTTPException(status_code=404, detail="vehicle not found")
    customer = models.Customer(**payload.model_dump())
    db.add(customer)
    db.commit()
    db.refresh(customer)
    return {"id": customer.id, "name": customer.name}


@app.get("/customers")
def list_customers(
    sort_by: Literal["name", "distance", "created_at", "updated_at"] = "name",
    sort_dir: Literal["asc", "desc"] = "asc",
    vehicle_id: int | None = None,
    all_catalogs: bool = False,
    db: Session = Depends(get_db),
) -> list[dict]:
    sort_column = {
        "name": models.Customer.name,
        "distance": models.Customer.distance_from_base_km,
        "created_at": models.Customer.created_at,
        "updated_at": models.Customer.updated_at,
    }[sort_by]
    order_expr = sort_column.asc() if sort_dir == "asc" else sort_column.desc()
    query = db.query(models.Customer).options(joinedload(models.Customer.vehicle))
    if not all_catalogs:
        if vehicle_id is None:
            query = query.filter(models.Customer.vehicle_id.is_(None))
        else:
            query = query.filter(models.Customer.vehicle_id == vehicle_id)
    rows = query.order_by(order_expr, models.Customer.id.asc()).all()
    return [serialize_customer(r) for r in rows]


@app.put("/customers/{customer_id}")
def update_customer(customer_id: int, payload: schemas.CustomerUpdate, db: Session = Depends(get_db)) -> dict:
    customer = db.get(models.Customer, customer_id)
    if not customer:
        raise HTTPException(status_code=404, detail="customer not found")
    if payload.vehicle_id and not db.get(models.Vehicle, payload.vehicle_id):
        raise HTTPException(status_code=404, detail="vehicle not found")
    for key, value in payload.model_dump().items():
        setattr(customer, key, value)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="customer update conflicts with existing data") from exc
    db.refresh(customer)
    return serialize_customer(customer)


@app.post("/customers/bulk-generation")
def bulk_update_customer_generation(payload: schemas.BulkCustomerGenerationUpdate, db: Session = Depends(get_db)) -> dict:
    rows = db.query(models.Customer).filter(models.Customer.id.in_(payload.ids)).all()
    if not rows:
        raise HTTPException(status_code=404, detail="no customers found")
    for row in rows:
        row.active_for_generation = payload.active_for_generation
    db.commit()
    return {
        "updated": len(rows),
        "active_for_generation": payload.active_for_generation,
    }


@app.delete("/customers/{customer_id}")
def delete_customer(customer_id: int, db: Session = Depends(get_db)) -> dict:
    customer = db.get(models.Customer, customer_id)
    if not customer:
        raise HTTPException(status_code=404, detail="customer not found")
    if db.query(models.Trip).filter(models.Trip.customer_id == customer_id).first():
        raise HTTPException(status_code=409, detail="customer is used in trips")
    db.delete(customer)
    db.commit()
    return {"deleted": True, "id": customer_id}


@app.post("/customers/import-csv")
async def import_customers_csv(
    file: UploadFile = File(...),
    name_column: str | None = Form(default=None),
    address_column: str | None = Form(default=None),
    distance_column: str | None = Form(default=None),
    db: Session = Depends(get_db),
) -> dict:
    if not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="only .csv files are supported")

    content = (await file.read()).decode("utf-8-sig")
    preview_line = content.splitlines()[0] if content.splitlines() else ""
    delimiter = ";" if preview_line.count(";") > preview_line.count(",") else ","
    reader = csv.DictReader(StringIO(content), delimiter=delimiter)
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="csv has no header")

    normalized_fields = {f.lower().strip(): f for f in reader.fieldnames}
    name_key = name_column if name_column in reader.fieldnames else normalized_fields.get("name") or normalized_fields.get("nazov")
    addr_key = address_column if address_column in reader.fieldnames else normalized_fields.get("address") or normalized_fields.get("adresa")
    if distance_column and distance_column in reader.fieldnames:
        dist_key = distance_column
    else:
        dist_key = (
            normalized_fields.get("distance_from_base_km")
            or normalized_fields.get("distance_km")
            or normalized_fields.get("vzdialenost_km")
        )
    if not name_key or not addr_key:
        raise HTTPException(status_code=400, detail="csv must contain name/nazov and address/adresa columns")

    settings_row = ensure_settings_row(db)
    created = 0
    updated = 0
    failed = 0
    errors: list[dict] = []

    for row_idx, row in enumerate(reader, start=2):
        name = (row.get(name_key) or "").strip()
        address = (row.get(addr_key) or "").strip()
        if not name or not address:
            failed += 1
            errors.append({"row": row_idx, "reason": "missing name or address"})
            continue
        distance_value = (row.get(dist_key) or "").strip() if dist_key else ""
        distance_km: float | None = None
        if distance_value:
            try:
                distance_km = float(distance_value.replace(",", "."))
            except ValueError:
                distance_km = None
                errors.append({"row": row_idx, "reason": f"invalid distance value: {distance_value}"})
        if distance_km is None and settings_row.company_base_address:
            try:
                estimate = distance_estimate(settings_row.company_base_address, address)
                distance_km = estimate["one_way_km"]
            except Exception:
                distance_km = None
        if distance_km is None:
            failed += 1
            errors.append({"row": row_idx, "reason": "distance missing and auto-distance failed"})
            continue

        existing = db.query(models.Customer).filter(models.Customer.name == name, models.Customer.address == address).first()
        if existing:
            existing.distance_from_base_km = distance_km
            updated += 1
        else:
            db.add(models.Customer(name=name, address=address, distance_from_base_km=distance_km))
            created += 1

    db.commit()
    return {
        "created": created,
        "updated": updated,
        "failed": failed,
        "errors": errors[:100],
        "detected_columns": reader.fieldnames,
        "used_mapping": {"name": name_key, "address": addr_key, "distance": dist_key},
    }


@app.post("/month-plans")
def create_month_plan(payload: schemas.MonthPlanCreate, db: Session = Depends(get_db)) -> dict:
    vehicle = db.get(models.Vehicle, payload.vehicle_id)
    driver = db.get(models.Driver, payload.driver_id)
    if not vehicle or not driver:
        raise HTTPException(status_code=404, detail="vehicle or driver not found")

    start_odometer_km = resolve_month_plan_start_odometer(
        db,
        payload.vehicle_id,
        payload.year,
        payload.month,
        payload.start_odometer_km,
    )
    if payload.end_odometer_km < start_odometer_km:
        raise HTTPException(status_code=400, detail="end_odometer_km must be >= start_odometer_km")

    month_plan = models.MonthPlan(**(payload.model_dump() | {"start_odometer_km": start_odometer_km}))
    db.add(month_plan)
    db.commit()
    db.refresh(month_plan)
    return {"id": month_plan.id, "year": month_plan.year, "month": month_plan.month}


def find_previous_month_plan(db: Session, vehicle_id: int, year: int, month: int) -> models.MonthPlan | None:
    previous_year = year - 1 if month == 1 else year
    previous_month = 12 if month == 1 else month - 1
    return (
        db.query(models.MonthPlan)
        .filter(
            models.MonthPlan.vehicle_id == vehicle_id,
            models.MonthPlan.year == previous_year,
            models.MonthPlan.month == previous_month,
        )
        .first()
    )


def resolve_month_plan_start_odometer(
    db: Session,
    vehicle_id: int,
    year: int,
    month: int,
    start_odometer_km: int | None,
) -> int:
    if start_odometer_km is not None:
        return start_odometer_km
    previous_plan = find_previous_month_plan(db, vehicle_id, year, month)
    if not previous_plan or previous_plan.end_odometer_km is None:
        raise HTTPException(
            status_code=400,
            detail="start_odometer_km is required when no previous month plan exists for this vehicle",
        )
    return previous_plan.end_odometer_km


@app.get("/month-plans")
def list_month_plans(
    year: int | None = None,
    month: int | None = None,
    vehicle_id: int | None = None,
    driver_id: int | None = None,
    db: Session = Depends(get_db),
) -> list[dict]:
    query = db.query(models.MonthPlan).options(
        joinedload(models.MonthPlan.vehicle),
        joinedload(models.MonthPlan.driver),
        joinedload(models.MonthPlan.trips),
    )
    if year is not None:
        query = query.filter(models.MonthPlan.year == year)
    if month is not None:
        query = query.filter(models.MonthPlan.month == month)
    if vehicle_id is not None:
        query = query.filter(models.MonthPlan.vehicle_id == vehicle_id)
    if driver_id is not None:
        query = query.filter(models.MonthPlan.driver_id == driver_id)
    rows = query.order_by(models.MonthPlan.id.desc()).all()
    return [serialize_month_plan(r) for r in rows]


@app.put("/month-plans/{month_plan_id}")
def update_month_plan(month_plan_id: int, payload: schemas.MonthPlanUpdate, db: Session = Depends(get_db)) -> dict:
    month_plan = db.get(models.MonthPlan, month_plan_id)
    if not month_plan:
        raise HTTPException(status_code=404, detail="month plan not found")
    if not db.get(models.Vehicle, payload.vehicle_id):
        raise HTTPException(status_code=404, detail="vehicle not found")
    if not db.get(models.Driver, payload.driver_id):
        raise HTTPException(status_code=404, detail="driver not found")

    start_odometer_km = resolve_month_plan_start_odometer(
        db,
        payload.vehicle_id,
        payload.year,
        payload.month,
        payload.start_odometer_km,
    )
    if payload.end_odometer_km < start_odometer_km:
        raise HTTPException(status_code=400, detail="end_odometer_km must be >= start_odometer_km")

    for key, value in (payload.model_dump() | {"start_odometer_km": start_odometer_km}).items():
        setattr(month_plan, key, value)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="month plan conflicts with existing period for vehicle") from exc
    db.refresh(month_plan)
    return serialize_month_plan(month_plan)


@app.delete("/month-plans/{month_plan_id}")
def delete_month_plan(month_plan_id: int, db: Session = Depends(get_db)) -> dict:
    month_plan = db.get(models.MonthPlan, month_plan_id)
    if not month_plan:
        raise HTTPException(status_code=404, detail="month plan not found")
    db.delete(month_plan)
    db.commit()
    return {"deleted": True, "id": month_plan_id}


def resolve_month_plan_for_refuel(payload: schemas.RefuelCreate, db: Session) -> models.MonthPlan:
    if payload.month_plan_id:
        month_plan = (
            db.query(models.MonthPlan)
            .options(joinedload(models.MonthPlan.vehicle))
            .filter(models.MonthPlan.id == payload.month_plan_id)
            .first()
        )
        if not month_plan:
            raise HTTPException(status_code=404, detail="month plan not found")
        return month_plan

    if not payload.vehicle_id:
        raise HTTPException(status_code=400, detail="month_plan_id or vehicle_id is required")

    month_plan = (
        db.query(models.MonthPlan)
        .options(joinedload(models.MonthPlan.vehicle))
        .filter(
            models.MonthPlan.vehicle_id == payload.vehicle_id,
            models.MonthPlan.year == payload.refuel_date.year,
            models.MonthPlan.month == payload.refuel_date.month,
        )
        .first()
    )
    if not month_plan:
        raise HTTPException(
            status_code=404,
            detail="no month plan found for selected vehicle and refuel date",
        )
    return month_plan


@app.post("/refuels")
def create_refuel(payload: schemas.RefuelCreate, db: Session = Depends(get_db)) -> dict:
    month_plan = resolve_month_plan_for_refuel(payload, db)
    if payload.liters > month_plan.vehicle.tank_capacity_l:
        raise HTTPException(
            status_code=400,
            detail=f"refuel liters exceed tank capacity ({month_plan.vehicle.tank_capacity_l} l)",
        )

    refuel_data = payload.model_dump(exclude={"vehicle_id"})
    refuel_data["month_plan_id"] = month_plan.id
    refuel = models.Refuel(**refuel_data)
    db.add(refuel)
    db.commit()
    db.refresh(refuel)
    return {"id": refuel.id, "liters": refuel.liters, "month_plan_id": refuel.month_plan_id}


@app.get("/month-plans/{month_plan_id}/refuels")
def list_refuels(month_plan_id: int, db: Session = Depends(get_db)) -> list[dict]:
    if not db.get(models.MonthPlan, month_plan_id):
        raise HTTPException(status_code=404, detail="month plan not found")

    rows = (
        db.query(models.Refuel)
        .filter(models.Refuel.month_plan_id == month_plan_id)
        .order_by(models.Refuel.refuel_date.asc(), models.Refuel.id.asc())
        .all()
    )
    return [serialize_refuel(r) for r in rows]


@app.put("/refuels/{refuel_id}")
def update_refuel(refuel_id: int, payload: schemas.RefuelUpdate, db: Session = Depends(get_db)) -> dict:
    refuel = db.get(models.Refuel, refuel_id)
    if not refuel:
        raise HTTPException(status_code=404, detail="refuel not found")
    month_plan = resolve_month_plan_for_refuel(payload, db)
    if payload.liters > month_plan.vehicle.tank_capacity_l:
        raise HTTPException(
            status_code=400,
            detail=f"refuel liters exceed tank capacity ({month_plan.vehicle.tank_capacity_l} l)",
        )

    update_data = payload.model_dump(exclude={"vehicle_id"})
    update_data["month_plan_id"] = month_plan.id
    for key, value in update_data.items():
        setattr(refuel, key, value)
    db.commit()
    db.refresh(refuel)
    return serialize_refuel(refuel)


@app.delete("/refuels/{refuel_id}")
def delete_refuel(refuel_id: int, db: Session = Depends(get_db)) -> dict:
    refuel = db.get(models.Refuel, refuel_id)
    if not refuel:
        raise HTTPException(status_code=404, detail="refuel not found")
    db.delete(refuel)
    db.commit()
    return {"deleted": True, "id": refuel_id}


@app.post("/trips")
def create_trip(payload: schemas.TripCreate, db: Session = Depends(get_db)) -> dict:
    month_plan = db.get(models.MonthPlan, payload.month_plan_id)
    if not month_plan:
        raise HTTPException(status_code=404, detail="month plan not found")
    if payload.customer_id and not db.get(models.Customer, payload.customer_id):
        raise HTTPException(status_code=404, detail="customer not found")
    validate_trip_range(month_plan, payload.trip_date, payload.trip_end_date)

    trip = models.Trip(**payload.model_dump(), generated=False)
    db.add(trip)
    db.commit()
    db.refresh(trip)
    return {"id": trip.id, "distance_km": trip.distance_km}


@app.get("/month-plans/{month_plan_id}/trips")
def list_trips(month_plan_id: int, db: Session = Depends(get_db)) -> list[dict]:
    if not db.get(models.MonthPlan, month_plan_id):
        raise HTTPException(status_code=404, detail="month plan not found")

    rows = (
        db.query(models.Trip)
        .options(joinedload(models.Trip.customer))
        .filter(models.Trip.month_plan_id == month_plan_id)
        .order_by(models.Trip.trip_date.asc(), models.Trip.id.asc())
        .all()
    )
    return [serialize_trip(r) for r in rows]


@app.get("/trips")
def list_all_trips(
    month_plan_id: int | None = None,
    year: int | None = None,
    month: int | None = None,
    vehicle_id: int | None = None,
    driver_id: int | None = None,
    mode: Literal["all", "manual", "generated", "private"] = "all",
    db: Session = Depends(get_db),
) -> list[dict]:
    query = (
        db.query(models.Trip)
        .join(models.MonthPlan, models.MonthPlan.id == models.Trip.month_plan_id)
        .options(joinedload(models.Trip.customer), joinedload(models.Trip.month_plan))
    )
    if month_plan_id is not None:
        query = query.filter(models.Trip.month_plan_id == month_plan_id)
    if year is not None:
        query = query.filter(models.MonthPlan.year == year)
    if month is not None:
        query = query.filter(models.MonthPlan.month == month)
    if vehicle_id is not None:
        query = query.filter(models.MonthPlan.vehicle_id == vehicle_id)
    if driver_id is not None:
        query = query.filter(models.MonthPlan.driver_id == driver_id)
    if mode == "manual":
        query = query.filter(models.Trip.generated.is_(False), models.Trip.is_private.is_(False))
    elif mode == "generated":
        query = query.filter(models.Trip.generated.is_(True), models.Trip.is_private.is_(False))
    elif mode == "private":
        query = query.filter(models.Trip.is_private.is_(True))

    rows = query.order_by(models.Trip.trip_date.asc(), models.Trip.id.asc()).all()
    return [serialize_trip(r) for r in rows]


@app.put("/trips/{trip_id}")
def update_trip(trip_id: int, payload: schemas.TripUpdate, db: Session = Depends(get_db)) -> dict:
    trip = db.get(models.Trip, trip_id)
    if not trip:
        raise HTTPException(status_code=404, detail="trip not found")
    month_plan = db.get(models.MonthPlan, payload.month_plan_id)
    if not month_plan:
        raise HTTPException(status_code=404, detail="month plan not found")
    if payload.customer_id and not db.get(models.Customer, payload.customer_id):
        raise HTTPException(status_code=404, detail="customer not found")
    validate_trip_range(month_plan, payload.trip_date, payload.trip_end_date)

    generated_flag = trip.generated
    for key, value in payload.model_dump().items():
        setattr(trip, key, value)
    trip.generated = generated_flag
    db.commit()
    db.refresh(trip)
    return serialize_trip(trip)


@app.delete("/trips/{trip_id}")
def delete_trip(trip_id: int, db: Session = Depends(get_db)) -> dict:
    trip = db.get(models.Trip, trip_id)
    if not trip:
        raise HTTPException(status_code=404, detail="trip not found")
    db.delete(trip)
    db.commit()
    return {"deleted": True, "id": trip_id}


@app.post("/vehicles/bulk-delete")
def bulk_delete_vehicles(payload: schemas.BulkDeleteRequest, db: Session = Depends(get_db)) -> dict:
    deleted = 0
    for entity_id in payload.ids:
        vehicle = db.get(models.Vehicle, entity_id)
        if not vehicle:
            continue
        if db.query(models.MonthPlan).filter(models.MonthPlan.vehicle_id == entity_id).first():
            continue
        db.query(models.Customer).filter(models.Customer.vehicle_id == entity_id).delete(synchronize_session=False)
        db.delete(vehicle)
        deleted += 1
    db.commit()
    return {"deleted": deleted}


@app.post("/drivers/bulk-delete")
def bulk_delete_drivers(payload: schemas.BulkDeleteRequest, db: Session = Depends(get_db)) -> dict:
    deleted = 0
    for entity_id in payload.ids:
        driver = db.get(models.Driver, entity_id)
        if not driver:
            continue
        if db.query(models.MonthPlan).filter(models.MonthPlan.driver_id == entity_id).first():
            continue
        db.delete(driver)
        deleted += 1
    db.commit()
    return {"deleted": deleted}


@app.post("/customers/bulk-delete")
def bulk_delete_customers(payload: schemas.BulkDeleteRequest, db: Session = Depends(get_db)) -> dict:
    deleted = 0
    for entity_id in payload.ids:
        customer = db.get(models.Customer, entity_id)
        if not customer:
            continue
        if db.query(models.Trip).filter(models.Trip.customer_id == entity_id).first():
            continue
        db.delete(customer)
        deleted += 1
    db.commit()
    return {"deleted": deleted}


@app.post("/month-plans/bulk-delete")
def bulk_delete_month_plans(payload: schemas.BulkDeleteRequest, db: Session = Depends(get_db)) -> dict:
    deleted = 0
    for entity_id in payload.ids:
        row = db.get(models.MonthPlan, entity_id)
        if not row:
            continue
        db.delete(row)
        deleted += 1
    db.commit()
    return {"deleted": deleted}


@app.post("/refuels/bulk-delete")
def bulk_delete_refuels(payload: schemas.BulkDeleteRequest, db: Session = Depends(get_db)) -> dict:
    deleted = 0
    for entity_id in payload.ids:
        row = db.get(models.Refuel, entity_id)
        if not row:
            continue
        db.delete(row)
        deleted += 1
    db.commit()
    return {"deleted": deleted}


@app.post("/trips/bulk-delete")
def bulk_delete_trips(payload: schemas.BulkDeleteRequest, db: Session = Depends(get_db)) -> dict:
    deleted = 0
    for entity_id in payload.ids:
        row = db.get(models.Trip, entity_id)
        if not row:
            continue
        db.delete(row)
        deleted += 1
    db.commit()
    return {"deleted": deleted}


@app.post("/month-plans/{month_plan_id}/generate", response_model=schemas.GenerateResponse)
def generate_month_trips(
    month_plan_id: int,
    payload: schemas.GenerateOptions | None = None,
    db: Session = Depends(get_db),
) -> schemas.GenerateResponse:
    month_plan = db.get(models.MonthPlan, month_plan_id)
    if not month_plan:
        raise HTTPException(status_code=404, detail="month plan not found")
    ensure_holidays_for_year(db, month_plan.year)
    if payload is not None:
        month_plan.private_km_enabled = payload.private_km_enabled
        month_plan.private_km_ratio_percent = payload.private_km_ratio_percent
        db.commit()
        db.refresh(month_plan)

    generated_trips, generated_km = generator.generate_missing_trips(db, month_plan)

    all_trips = db.query(models.Trip).filter(models.Trip.month_plan_id == month_plan_id).all()
    target_km = month_plan.end_odometer_km - month_plan.start_odometer_km
    hidden_private_km = round(sum(t.distance_km for t in all_trips if t.is_private), 1)
    service_target_km = generator.plan_service_target_km(month_plan)
    recorded_service_km = round(sum(t.distance_km for t in all_trips if not t.is_private), 1)
    total_km = round(recorded_service_km + hidden_private_km, 1)
    total_refueled = round(sum(r.liters for r in month_plan.refuels), 1)
    estimated_fuel = round((target_km * month_plan.vehicle.expected_consumption_l_per_100km) / 100.0, 1)

    warning = None
    delta = round(total_refueled - estimated_fuel, 1)
    if abs(delta) > 10:
        warning = (
            "Rozdiel medzi odhadovanou spotrebou a tankovanim je vacsi ako 10 l. "
            "Skontroluj manualne jazdy alebo spotrebu vozidla."
        )

    return schemas.GenerateResponse(
        generated_trips=generated_trips,
        generated_km=generated_km,
        target_km=target_km,
        service_target_km=service_target_km,
        hidden_private_km=hidden_private_km,
        recorded_service_km=recorded_service_km,
        total_km_including_private=total_km,
        total_trips_after_generation=len(all_trips),
        estimated_fuel_l=estimated_fuel,
        refueled_l=total_refueled,
        warning=warning,
    )


@app.get("/month-plans/{month_plan_id}/report", response_model=schemas.MonthReport)
def get_month_report(month_plan_id: int, db: Session = Depends(get_db)) -> schemas.MonthReport:
    month_plan = db.get(models.MonthPlan, month_plan_id)
    if not month_plan:
        raise HTTPException(status_code=404, detail="month plan not found")

    all_trips = db.query(models.Trip).filter(models.Trip.month_plan_id == month_plan_id).all()
    recorded_service_km = round(sum(t.distance_km for t in all_trips if not t.is_private), 1)
    target_km = month_plan.end_odometer_km - month_plan.start_odometer_km
    hidden_private_km = round(sum(t.distance_km for t in all_trips if t.is_private), 1)
    service_target_km = generator.plan_service_target_km(month_plan)
    total_km = round(recorded_service_km + hidden_private_km, 1)
    refueled_l = round(sum(r.liters for r in month_plan.refuels), 1)
    estimated_fuel_l = round((target_km * month_plan.vehicle.expected_consumption_l_per_100km) / 100.0, 1)
    avg_consumption = round((refueled_l / total_km) * 100.0, 2) if total_km > 0 else None

    return schemas.MonthReport(
        month_plan_id=month_plan.id,
        target_km=target_km,
        total_km=total_km,
        recorded_service_km=recorded_service_km,
        hidden_private_km=hidden_private_km,
        service_target_km=service_target_km,
        refueled_l=refueled_l,
        estimated_fuel_l=estimated_fuel_l,
        fuel_difference_l=round(refueled_l - estimated_fuel_l, 1),
        average_consumption_l_per_100km=avg_consumption,
        trips_count=len(all_trips),
    )


@app.get("/month-plans/{month_plan_id}/trips.csv")
def export_trips_csv(
    month_plan_id: int,
    mode: Literal["all", "manual", "generated", "private"] = "all",
    db: Session = Depends(get_db),
) -> Response:
    month_plan = (
        db.query(models.MonthPlan)
        .options(joinedload(models.MonthPlan.vehicle), joinedload(models.MonthPlan.driver))
        .filter(models.MonthPlan.id == month_plan_id)
        .first()
    )
    if not month_plan:
        raise HTTPException(status_code=404, detail="month plan not found")

    query = db.query(models.Trip).options(joinedload(models.Trip.customer)).filter(models.Trip.month_plan_id == month_plan_id)
    if mode == "manual":
        query = query.filter(models.Trip.generated.is_(False), models.Trip.is_private.is_(False))
    elif mode == "generated":
        query = query.filter(models.Trip.generated.is_(True), models.Trip.is_private.is_(False))
    elif mode == "private":
        query = query.filter(models.Trip.is_private.is_(True))
    rows = query.order_by(models.Trip.trip_date.asc(), models.Trip.id.asc()).all()
    export_rows = build_export_rows_for_month_plan(month_plan, rows)
    csv_content = render_trip_export_csv(export_rows)

    filename = f"month_plan_{month_plan_id}_trips_{mode}.csv"
    return Response(
        content=csv_content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/month-plans/{month_plan_id}/trips.xlsx")
def export_trips_xlsx(
    month_plan_id: int,
    mode: Literal["all", "manual", "generated", "private"] = "all",
    db: Session = Depends(get_db),
) -> Response:
    target_month_plan = (
        db.query(models.MonthPlan)
        .options(joinedload(models.MonthPlan.vehicle), joinedload(models.MonthPlan.driver))
        .filter(models.MonthPlan.id == month_plan_id)
        .first()
    )
    if not target_month_plan:
        raise HTTPException(status_code=404, detail="month plan not found")

    settings_row = ensure_settings_row(db)
    workbook = load_workbook(TRIPS_TEMPLATE_XLSX)
    year_plans = (
        db.query(models.MonthPlan)
        .options(joinedload(models.MonthPlan.vehicle), joinedload(models.MonthPlan.driver))
        .filter(
            models.MonthPlan.vehicle_id == target_month_plan.vehicle_id,
            models.MonthPlan.year == target_month_plan.year,
        )
        .order_by(models.MonthPlan.month.asc())
        .all()
    )
    for plan in year_plans:
        query = db.query(models.Trip).options(joinedload(models.Trip.customer)).filter(models.Trip.month_plan_id == plan.id)
        if mode == "manual":
            query = query.filter(models.Trip.generated.is_(False), models.Trip.is_private.is_(False))
        elif mode == "generated":
            query = query.filter(models.Trip.generated.is_(True), models.Trip.is_private.is_(False))
        elif mode == "private":
            query = query.filter(models.Trip.is_private.is_(True))
        rows = query.order_by(models.Trip.trip_date.asc(), models.Trip.id.asc()).all()
        sheet = _resolve_template_sheet(workbook, plan.month, plan.year)
        _fill_template_month_sheet(sheet, plan, rows, settings_row.company_name)

    stream = BytesIO()
    workbook.save(stream)
    xlsx_content = stream.getvalue()

    filename = f"kniha_jazd_{target_month_plan.vehicle.plate_number}_{target_month_plan.year}_{mode}.xlsx"
    return Response(
        content=xlsx_content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def query_filtered_trips_for_export(
    month_plan_id: int | None = None,
    year: int | None = None,
    month: int | None = None,
    vehicle_id: int | None = None,
    driver_id: int | None = None,
    mode: Literal["all", "manual", "generated", "private"] = "all",
    db: Session | None = None,
) -> list[models.Trip]:
    if db is None:
        raise ValueError("db session is required")
    query = (
        db.query(models.Trip)
        .join(models.MonthPlan, models.MonthPlan.id == models.Trip.month_plan_id)
        .options(
            joinedload(models.Trip.customer),
            joinedload(models.Trip.month_plan).joinedload(models.MonthPlan.vehicle),
            joinedload(models.Trip.month_plan).joinedload(models.MonthPlan.driver),
        )
    )
    if month_plan_id is not None:
        query = query.filter(models.Trip.month_plan_id == month_plan_id)
    if year is not None:
        query = query.filter(models.MonthPlan.year == year)
    if month is not None:
        query = query.filter(models.MonthPlan.month == month)
    if vehicle_id is not None:
        query = query.filter(models.MonthPlan.vehicle_id == vehicle_id)
    if driver_id is not None:
        query = query.filter(models.MonthPlan.driver_id == driver_id)
    if mode == "manual":
        query = query.filter(models.Trip.generated.is_(False), models.Trip.is_private.is_(False))
    elif mode == "generated":
        query = query.filter(models.Trip.generated.is_(True), models.Trip.is_private.is_(False))
    elif mode == "private":
        query = query.filter(models.Trip.is_private.is_(True))

    return query.order_by(models.MonthPlan.year.asc(), models.MonthPlan.month.asc(), models.Trip.trip_date.asc(), models.Trip.id.asc()).all()


@app.get("/trips.csv")
def export_filtered_trips_csv(
    month_plan_id: int | None = None,
    year: int | None = None,
    month: int | None = None,
    vehicle_id: int | None = None,
    driver_id: int | None = None,
    mode: Literal["all", "manual", "generated", "private"] = "all",
    db: Session = Depends(get_db),
) -> Response:
    trips = query_filtered_trips_for_export(
        month_plan_id=month_plan_id,
        year=year,
        month=month,
        vehicle_id=vehicle_id,
        driver_id=driver_id,
        mode=mode,
        db=db,
    )
    export_rows = build_export_rows_for_mixed_trips(trips)
    csv_content = render_trip_export_csv(export_rows)

    year_part = year if year is not None else "all"
    month_part = month if month is not None else "all"
    plan_part = month_plan_id if month_plan_id is not None else "all"
    filename = f"trips_plan{plan_part}_y{year_part}_m{month_part}_{mode}.csv"
    return Response(
        content=csv_content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/trips.xlsx")
def export_filtered_trips_xlsx(
    month_plan_id: int | None = None,
    year: int | None = None,
    month: int | None = None,
    vehicle_id: int | None = None,
    driver_id: int | None = None,
    mode: Literal["all", "manual", "generated", "private"] = "all",
    db: Session = Depends(get_db),
) -> Response:
    trips = query_filtered_trips_for_export(
        month_plan_id=month_plan_id,
        year=year,
        month=month,
        vehicle_id=vehicle_id,
        driver_id=driver_id,
        mode=mode,
        db=db,
    )
    export_rows = build_export_rows_for_mixed_trips(trips)
    xlsx_content = render_trip_export_xlsx(export_rows)

    year_part = year if year is not None else "all"
    month_part = month if month is not None else "all"
    plan_part = month_plan_id if month_plan_id is not None else "all"
    filename = f"trips_plan{plan_part}_y{year_part}_m{month_part}_{mode}.xlsx"
    return Response(
        content=xlsx_content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/month-plans/{month_plan_id}/full-export.xlsx")
def export_month_plan_full_xlsx(month_plan_id: int, db: Session = Depends(get_db)) -> Response:
    target_month_plan = (
        db.query(models.MonthPlan)
        .options(joinedload(models.MonthPlan.vehicle), joinedload(models.MonthPlan.driver))
        .filter(models.MonthPlan.id == month_plan_id)
        .first()
    )
    if not target_month_plan:
        raise HTTPException(status_code=404, detail="month plan not found")

    year_plans = (
        db.query(models.MonthPlan)
        .options(joinedload(models.MonthPlan.vehicle), joinedload(models.MonthPlan.driver))
        .filter(
            models.MonthPlan.vehicle_id == target_month_plan.vehicle_id,
            models.MonthPlan.year == target_month_plan.year,
        )
        .order_by(models.MonthPlan.month.asc())
        .all()
    )
    year_plan_ids = [p.id for p in year_plans]

    settings_row = ensure_settings_row(db)
    wb = load_workbook(TRIPS_TEMPLATE_XLSX)
    for plan in year_plans:
        trips = (
            db.query(models.Trip)
            .options(joinedload(models.Trip.customer))
            .filter(models.Trip.month_plan_id == plan.id)
            .order_by(models.Trip.trip_date.asc(), models.Trip.id.asc())
            .all()
        )
        sheet = _resolve_template_sheet(wb, plan.month, plan.year)
        _fill_template_month_sheet(sheet, plan, trips, settings_row.company_name)

    refuels = (
        db.query(models.Refuel)
        .join(models.MonthPlan, models.MonthPlan.id == models.Refuel.month_plan_id)
        .options(joinedload(models.Refuel.month_plan))
        .filter(models.Refuel.month_plan_id.in_(year_plan_ids))
        .order_by(models.MonthPlan.month.asc(), models.Refuel.refuel_date.asc(), models.Refuel.id.asc())
        .all()
    )

    refuels_sheet = wb.create_sheet(title="Tankovania")
    refuels_sheet.append(["Mesiac", "Datum", "Litre", "Cena EUR", "Mesto", "Zahranicie", "Tachometer km"])
    for row in refuels:
        refuels_sheet.append(
            [
                MONTH_NAMES_SK.get(row.month_plan.month, str(row.month_plan.month)).capitalize(),
                format_date_sk(row.refuel_date.isoformat()),
                row.liters,
                row.total_price_eur,
                row.location_city,
                "ano" if row.is_foreign else "nie",
                row.odometer_km,
            ]
        )

    stream = BytesIO()
    wb.save(stream)
    filename = f"kniha_jazd_{target_month_plan.vehicle.plate_number}_{target_month_plan.year}_full_export.xlsx"
    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


@app.get("/month-plans/{month_plan_id}/report.csv")
def export_report_csv(month_plan_id: int, db: Session = Depends(get_db)) -> Response:
    month_plan = db.get(models.MonthPlan, month_plan_id)
    if not month_plan:
        raise HTTPException(status_code=404, detail="month plan not found")

    all_trips = db.query(models.Trip).filter(models.Trip.month_plan_id == month_plan_id).all()
    recorded_service_km = round(sum(t.distance_km for t in all_trips if not t.is_private), 1)
    target_km = month_plan.end_odometer_km - month_plan.start_odometer_km
    hidden_private_km = round(sum(t.distance_km for t in all_trips if t.is_private), 1)
    service_target_km = generator.plan_service_target_km(month_plan)
    total_km = round(recorded_service_km + hidden_private_km, 1)
    refueled_l = round(sum(r.liters for r in month_plan.refuels), 1)
    estimated_fuel_l = round((target_km * month_plan.vehicle.expected_consumption_l_per_100km) / 100.0, 1)
    average_consumption = round((refueled_l / total_km) * 100.0, 2) if total_km > 0 else ""

    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "month_plan_id",
            "target_km",
            "service_target_km",
            "recorded_service_km",
            "hidden_private_km",
            "total_km",
            "refueled_l",
            "estimated_fuel_l",
            "fuel_difference_l",
            "average_consumption_l_per_100km",
        ]
    )
    writer.writerow(
        [
            month_plan.id,
            target_km,
            service_target_km,
            recorded_service_km,
            hidden_private_km,
            total_km,
            refueled_l,
            estimated_fuel_l,
            round(refueled_l - estimated_fuel_l, 1),
            average_consumption,
        ]
    )

    filename = f"month_plan_{month_plan_id}_report.csv"
    return Response(
        content=buffer.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
